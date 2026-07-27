import os
import re
from flask import (
    Blueprint,
    render_template,
    request,
    redirect,
    url_for,
    flash
)
from flask import send_from_directory
from flask import send_file
from flask_login import login_required
from openpyxl import load_workbook
from werkzeug.utils import secure_filename

from extensions import db
from models.member_index import MemberIndex
from utils.auth import admin_required

member_indexes_bp = Blueprint(
    "member_indexes",
    __name__,
    url_prefix="/member-indexes"
)


@member_indexes_bp.route("/")
@login_required
@admin_required
def index():

    indexes = MemberIndex.query.order_by(
        MemberIndex.student_id
    ).all()

    return render_template(
        "member_indexes/index.html",
        indexes=indexes
    )


@member_indexes_bp.route("/add", methods=["GET", "POST"])
@login_required
@admin_required
def add():

    if request.method == "POST":

        student_id = request.form["student_id"].strip().upper()

        exists = MemberIndex.query.filter_by(
            student_id=student_id
        ).first()

        if exists:

            flash(
                "Student Index Number already exists.",
                "warning"
            )

            return redirect(
                url_for("member_indexes.add")
            )

        db.session.add(
            MemberIndex(student_id=student_id)
        )

        db.session.commit()

        flash(
            "Student Index Number added successfully.",
            "success"
        )

        return redirect(
            url_for("member_indexes.index")
        )

    return render_template(
        "member_indexes/add.html"
    )

# ==========================================
# Edit Student Index
# ==========================================
@member_indexes_bp.route("/<int:index_id>/edit", methods=["GET", "POST"])
@login_required
@admin_required
def edit(index_id):

    index = MemberIndex.query.get_or_404(index_id)

    if request.method == "POST":

        student_id = request.form["student_id"].strip().upper()

        exists = MemberIndex.query.filter(
            MemberIndex.student_id == student_id,
            MemberIndex.id != index.id
        ).first()

        if exists:

            flash(
                "Student Index Number already exists.",
                "warning"
            )

            return redirect(
                url_for(
                    "member_indexes.edit",
                    index_id=index.id
                )
            )

        index.student_id = student_id

        db.session.commit()

        flash(
            "Student Index updated successfully.",
            "success"
        )

        return redirect(
            url_for("member_indexes.index")
        )

    return render_template(
        "member_indexes/edit.html",
        index=index
    )


# ==========================================
# Delete Student Index
# ==========================================
@member_indexes_bp.route("/<int:index_id>/delete", methods=["GET", "POST"])
@login_required
@admin_required
def delete(index_id):

    index = MemberIndex.query.get_or_404(index_id)

    if request.method == "POST":

        # If linked to a user, clear the relationship first
        index.used = False
        index.used_by = None
        index.used_at = None

        db.session.delete(index)
        db.session.commit()

        flash(
            "Student Index deleted successfully.",
            "success"
        )

        return redirect(
            url_for("member_indexes.index")
        )

    return render_template(
        "member_indexes/delete.html",
        index=index
    )


# ==========================================
# Bulk Delete Student Indexes
# ==========================================
@member_indexes_bp.route("/bulk-delete", methods=["POST"])
@login_required
@admin_required
def bulk_delete():

    selected_ids = request.form.getlist("selected_ids")

    if not selected_ids:
        flash("No Student IDs selected.", "warning")
        return redirect(url_for("member_indexes.index"))

    indexes = MemberIndex.query.filter(
        MemberIndex.id.in_(selected_ids)
    ).all()

    for index in indexes:
        db.session.delete(index)

    db.session.commit()

    flash(
        f"{len(indexes)} Student ID(s) deleted successfully.",
        "success"
    )

    return redirect(url_for("member_indexes.index"))



@member_indexes_bp.route("/import", methods=["GET", "POST"])
@login_required
@admin_required
def import_indexes():

    if request.method == "POST":

        file = request.files.get("excel_file")

        if not file or file.filename == "":
            flash("Please select an Excel file.", "danger")
            return redirect(request.url)

        if not file.filename.lower().endswith(".xlsx"):
            flash("Only Excel (.xlsx) files are allowed.", "danger")
            return redirect(request.url)

        upload_folder = "uploads"
        os.makedirs(upload_folder, exist_ok=True)

        filename = secure_filename(file.filename)
        filepath = os.path.join(upload_folder, filename)

        file.save(filepath)

        workbook = load_workbook(filepath)
        sheet = workbook.active

        total_rows = 0
        imported = 0
        db_duplicates = 0
        excel_duplicates = 0
        empty_rows = 0

        seen = set()

        for row in sheet.iter_rows(min_row=2, values_only=True):

            total_rows += 1

            student_id = None

            # Search every column for a valid 10-digit Student Index Number
            for value in row:

                if value is None:
                    continue

                value = str(value).strip()
                value = value.replace(" ", "")

                # Only accept exactly 10 digits
                if re.fullmatch(r"\d{10}", value):
                    student_id = value
                    break

            # No valid Student Index Number found
            if student_id is None:
                empty_rows += 1
                continue
                # Duplicate inside Excel
            if student_id in seen:

                excel_duplicates += 1
                continue

            seen.add(student_id)

            # Duplicate in database
            exists = MemberIndex.query.filter_by(
                student_id=student_id
            ).first()

            if exists:

                db_duplicates += 1
                continue

            db.session.add(
                MemberIndex(
                    student_id=student_id
                )
            )

            imported += 1

        db.session.commit()

        flash(
            f"""
Import Completed

Total Rows : {total_rows}

Imported : {imported}

Database Duplicates : {db_duplicates}

Excel Duplicates : {excel_duplicates}

Empty Rows : {empty_rows}
""",
            "success"
        )

        return redirect(
            url_for("member_indexes.index")
        )

    return render_template(
        "member_indexes/import.html"
    )
@member_indexes_bp.route("/download-sample")
@login_required
@admin_required
def download_sample():

    return send_from_directory(
        "static/templates",
        "ESA_Student_Index_Template.xlsx",
        as_attachment=True
    )