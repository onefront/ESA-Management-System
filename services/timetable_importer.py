import os
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import re
class TimetableImporter:

    def __init__(self, filepath):
        self.filepath = filepath

    def normalize_level(self, level):

        if level is None:
            return ""

        level = str(level).strip().upper()

        mapping = {
            "I": "100",
            "II": "200",
            "III": "300",
            "IV": "400"
        }

        return mapping.get(level, level)

    def normalize_programme(self, programme):

        if programme is None:
            return ""

        return str(programme).strip().upper()

    import re

    def extract_class_details(self, class_text):

        if not class_text:
            return "", ""

        text = str(class_text).upper().strip()

        # -------------------------
        # LEVEL
        # -------------------------
        level = ""

        if re.search(r"\bIV\b", text):
            level = "400"
        elif re.search(r"\bIII\b", text):
            level = "300"
        elif re.search(r"\bII\b", text):
            level = "200"
        elif re.search(r"\bI\b", text):
            level = "100"

        # -------------------------
        # Remove prefixes
        # -------------------------
        cleaned = text

        replacements = [
            "4-YR.",
            "2-YR.",
            "4-YR",
            "2-YR",
            "BSC.",
            "BSC",
            "BBA.",
            "BBA",
            "DIP.",
            "DIP",
        ]

        for item in replacements:
            cleaned = cleaned.replace(item, " ")

        cleaned = re.sub(r"\bIV\b", "", cleaned)
        cleaned = re.sub(r"\bIII\b", "", cleaned)
        cleaned = re.sub(r"\bII\b", "", cleaned)
        cleaned = re.sub(r"\bI\b", "", cleaned)

        cleaned = " ".join(cleaned.split())

        # -------------------------
        # Special programmes
        # -------------------------
        if "MKT & ENT" in cleaned:
            return "MKT & ENT", level

        if "B/F" in cleaned:
            return "B/F", level

        if "B&F" in cleaned:
            return "B&F", level

        # -------------------------
        # Extract last meaningful word
        # -------------------------
        words = cleaned.split()

        ignore = {
            "B",
            "YR",
            "DIP",
            "BSC",
            "BBA"
        }

        words = [w for w in words if w not in ignore]

        if words:
            return words[-1], level

        return "", level


    def fill_merged_cells(self, sheet):

        merged_ranges = list(sheet.merged_cells.ranges)

        for merged_range in merged_ranges:

            min_col = merged_range.min_col
            min_row = merged_range.min_row
            max_col = merged_range.max_col
            max_row = merged_range.max_row

            value = sheet.cell(min_row, min_col).value

            sheet.unmerge_cells(str(merged_range))

            for row in range(min_row, max_row + 1):
                for col in range(min_col, max_col + 1):
                    sheet.cell(row, col).value = value

    def import_excel(self):

        workbook = load_workbook(
            self.filepath,
            data_only=True
        )

        sheet = workbook.active
        print("=" * 50)
        print("A1 =", sheet["A1"].value)
        print("A2 =", sheet["A2"].value)
        print("A5 =", sheet["A5"].value)
        print("A6 =", sheet["A6"].value)
        print("=" * 50)
        print("=" * 120)

        for i, row in enumerate(sheet.iter_rows(values_only=True), start=1):

            print(i, row)

            if i == 80:
                break

        print("=" * 120)
        self.fill_merged_cells(sheet)
        for row in sheet.iter_rows(min_row=1, max_row=40, values_only=True):
            print(row)
        records = []

        # Skip header row
        for row in sheet.iter_rows(min_row=2, values_only=True):

            if all(cell is None for cell in row):
                continue

            programme, level = self.extract_class_details(row[4])

            records.append({
                "exam_date": row[0],
                "day": row[1],
                "start_time": row[2],
                "end_time": row[3],
                "programme": programme,
                "level": level,
                "session": row[6],
                "course_code": row[7],
                "course_title": row[8],
                "venue": row[9],
                "examiner": row[10]
            })

        return records