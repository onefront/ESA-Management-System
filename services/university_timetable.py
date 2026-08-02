from openpyxl import load_workbook
from services.timetable_importer import TimetableImporter
from datetime import datetime, timedelta
class UniversityTimetableImporter:

    def __init__(self, filepath):
        self.filepath = filepath
        self.helper = TimetableImporter(filepath)

    def import_excel(self):

        workbook = load_workbook(
            self.filepath,
            data_only=True
        )

        sheet = workbook.active

        # Expand merged cells
        self.helper.fill_merged_cells(sheet)

        records = []

        started = False

        current_day = ""
        current_date = None
        current_time = None

        last_venue = ""
        last_examiner = ""
        last_course_code = ""
        last_course_title = ""
        for row in sheet.iter_rows(values_only=True):

            # Skip completely empty rows
            if not any(row):
                continue

            # Wait until timetable header
            if not started:

                if row[0] == "DAY/DATE":
                    started = True

                continue

            # -----------------------------
            # DATE / DAY
            # -----------------------------
            if row[0]:

                value = str(row[0]).strip()

                # Actual date
                if "/" in value:
                    current_date = datetime.strptime(
                        value,
                        "%d/%m/%Y"
                    ).date()
                    current_day = ""

                # Day written vertically
                elif value in ["S", "A", "T", "U", "R", "D", "Y"]:

                    current_day += value

                    if current_day == "SATURDAY":
                        current_day = "Saturday"

                    elif current_day == "SUNDAY":
                        current_day = "Sunday"

            # -----------------------------
            # TIME
            # -----------------------------
            if row[1]:
                current_time = datetime.strptime(
                    str(row[1]).strip().upper(),
                    "%I:%M%p"
                ).time()

            # -----------------------------
            # CLASS
            # -----------------------------
            if not row[2]:
                continue

            programme, level = self.helper.extract_class_details(row[2])

            if current_date is None:
                continue
            if row[6]:
                last_venue = row[6]

            if row[7]:
                last_examiner = row[7]
            if row[3]:
                last_course_code = row[3]

            if row[4]:
                last_course_title = row[4]
            # Calculate end time (3 hours after start time)
            if current_time:
                end_time = (
                        datetime.combine(
                            datetime.today(),
                            current_time
                        ) + timedelta(hours=3)
                ).time()
            else:
                end_time = None

            records.append({

                "day": current_day,
                "exam_date": current_date,

                "start_time": current_time,
                "end_time": end_time,

                "programme": programme,
                "level": level,
                "session": "Weekend",

                "class": row[2],

                "course_code": last_course_code,
                "course_title": last_course_title,

                "students": row[5],

                "venue": last_venue,
                "examiner": last_examiner


            })
        print("TOTAL RECORDS:", len(records))

        for r in records[:20]:
            print(r)

        print("=" * 80)

        return records