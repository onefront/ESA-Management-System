import os

UPLOAD_FOLDER = "static/uploads/imports"

os.makedirs(
    UPLOAD_FOLDER,
    exist_ok=True
)
from openpyxl import load_workbook

from flask import (
    render_template,
    request,
    redirect,
    flash,
    url_for,
    send_file
)

from werkzeug.utils import secure_filename

from extensions import db

from models.programme import Programme
from models.faculty import Faculty

from flask_login import login_required

from utils.auth import roles_required

from flask import Blueprint

data_management_bp = Blueprint(
    "data_management",
    __name__,
    url_prefix="/data-management"
)




@data_management_bp.route("/")
def dashboard():
    pass



@data_management_bp.route(
    "/import-programmes",
    methods=["GET", "POST"]
)
@login_required
@roles_required("Administrator")
def import_programmes():

    if request.method == "POST":

        file = request.files.get("file")

        if not file or file.filename == "":
            flash("Please choose an Excel file.", "danger")
            return redirect(request.url)

        filename = secure_filename(file.filename)

        filepath = os.path.join(
            UPLOAD_FOLDER,
            filename
        )

        file.save(filepath)

        workbook = load_workbook(filepath)
        sheet = workbook.active

        imported = 0
        updated = 0
        skipped = 0
        errors = 0

        for row in sheet.iter_rows(min_row=2, values_only=True):

            if not row:
                continue

            programme_name = str(row[0] or "").strip()
            programme_code = str(row[1] or "").strip().upper()
            faculty_name = str(row[2] or "").strip()
            status = str(row[3] or "Active").strip()

            faculty = Faculty.query.filter_by(
                faculty_name=faculty_name
            ).first()

            if not faculty:
                errors += 1
                continue

            programme = Programme.query.filter_by(
                programme_name=programme_name
            ).first()

            if programme:

                changed = False

                if programme.programme_code != programme_code:
                    programme.programme_code = programme_code
                    changed = True

                if programme.faculty_id != faculty.id:
                    programme.faculty_id = faculty.id
                    changed = True

                if programme.status != status:
                    programme.status = status
                    changed = True

                if changed:
                    updated += 1
                else:
                    skipped += 1

            else:

                programme = Programme(
                    programme_name=programme_name,
                    programme_code=programme_code,
                    faculty_id=faculty.id,
                    status=status
                )

                db.session.add(programme)
                imported += 1

        db.session.commit()

        flash(
            f"Import completed. Imported: {imported}, Updated: {updated}, Skipped: {skipped}, Errors: {errors}",
            "success"
        )

        return redirect(
            url_for("data_management.import_programmes")
        )

    return render_template(
        "data_management/import_programmes.html"
    )


@data_management_bp.route("/download-template")
@login_required
@roles_required("Administrator")
def download_programme_template():

    return send_file(
        "static/templates/programme_import_template.xlsx",
        as_attachment=True
    )