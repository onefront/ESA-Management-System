from openpyxl import Workbook
from openpyxl.styles import Font

wb = Workbook()
ws = wb.active

ws.title = "Programmes"

headers = [
    "Programme Name",
    "Programme Code",
    "Faculty",
    "Status"
]

for col, header in enumerate(headers, start=1):
    cell = ws.cell(row=1, column=col)
    cell.value = header
    cell.font = Font(bold=True)

programmes = [
    ["B.Sc. Information Technology Education", "ITE", "INFORMATION TECHNOLOGY EDUCATION", "Active"],
    ["B.Sc. Information Technology", "IT", "INFORMATION TECHNOLOGY EDUCATION", "Active"],
    ["B.Sc. Cyber Security and Digital Forensics", "CSDF", "INFORMATION TECHNOLOGY EDUCATION", "Active"],
    ["B.Ed. Information Technology", "BED-IT", "INFORMATION TECHNOLOGY EDUCATION", "Active"],
    ["B.Ed. Computing with Artificial Intelligence (AI)", "BED-AI", "INFORMATION TECHNOLOGY EDUCATION", "Active"],
    ["B.Ed. Computing with Internet of Things (IoT)", "BED-IOT", "INFORMATION TECHNOLOGY EDUCATION", "Active"],
]

row = 2

for programme in programmes:

    for col, value in enumerate(programme, start=1):

        ws.cell(row=row, column=col).value = value

    row += 1

for column_cells in ws.columns:

    length = max(len(str(cell.value or "")) for cell in column_cells)

    ws.column_dimensions[column_cells[0].column_letter].width = length + 5

wb.save("static/templates/programme_import_template.xlsx")

print("Programme template created successfully.")