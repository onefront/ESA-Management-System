from flask import (
    Blueprint,
    send_file,
    render_template,
    request,
    redirect,
    url_for,
    flash
)

from openpyxl import Workbook, load_workbook

from models.member import Member
from extensions import db

member_import_bp = Blueprint("member_import", __name__)


# ==========================================
# Download Excel Template
# ==========================================
@member_import_bp.route("/members/template")
def download_template():

    wb = Workbook()
    ws = wb.active
    ws.title = "ESA Members"

    headers = [
        "ESA ID",
        "Student ID",
        "First Name",
        "Last Name",
        "Phone",
        "Email",
        "Programme",
        "Level",
        "Session"
    ]

    ws.append(headers)

    ws.append([
        "ESA001",
        "2026/ITE/001",
        "Owusu",
        "Issah",
        "0241234567",
        "example@email.com",
        "Information Technology",
        "300",
        "Weekend"
    ])

    filename = "ESA_Member_Template.xlsx"

    wb.save(filename)

    return send_file(
        filename,
        as_attachment=True
    )


# ==========================================
# Import Members
# ==========================================
@member_import_bp.route("/members/import", methods=["GET", "POST"])
def import_members():

    if request.method == "POST":

        file = request.files.get("excel_file")

        if not file:
            flash("Please select an Excel file.", "danger")
            return redirect(request.url)

        workbook = load_workbook(file)
        sheet = workbook.active

        imported = 0
        skipped = 0

        for row in sheet.iter_rows(min_row=2, values_only=True):

            # Skip empty rows
            if not row[1]:
                continue

            # Check duplicate Student ID or ESA ID
            existing = Member.query.filter(
                (Member.student_id == row[1]) |
                (Member.esa_id == row[0])
            ).first()

            if existing:
                skipped += 1
                continue

            member = Member(
                esa_id=row[0],
                student_id=row[1],
                first_name=row[2],
                last_name=row[3],
                phone=row[4],
                email=row[5],
                programme=row[6],
                level=row[7],
                session=row[8]
            )

            db.session.add(member)
            imported += 1

        db.session.commit()

        flash(
            f"Imported: {imported} member(s). Skipped: {skipped} duplicate(s).",
            "success"
        )

        return redirect(url_for("members.members"))

    return render_template("import.html")